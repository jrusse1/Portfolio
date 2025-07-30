import openpyxl

# Q1
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
print('Number of rows found in sheet: ', sheet.max_row)
print('Number of columns found in sheet: ', sheet.max_column)

# Q2
total = 0
for i in range(2, sheet.max_row):
    cell_value = sheet.cell(row=i, column=4).value
    total += cell_value
print('Total in column D: ', total)

# Q3
state_counter = {}
def addone(state):
     if state in state_counter:
           state_counter[state] += 1
     else:
           state_counter[state] = 1
for state in range (2, sheet.max_row):
    cell_value = sheet.cell(row=state, column=2).value
    addone(cell_value)
print(state_counter)
